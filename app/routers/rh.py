import io
from datetime import datetime, timedelta, timezone
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File as FastAPIFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import func
from sqlalchemy.orm import Session

from sqlalchemy.orm.attributes import flag_modified

from app.config import settings
from app.database import get_db
from app.models.demande import Demande
from app.models.employee import Employe
from app.models.rh import RH
from app.models.parametrage import Parametrage
from app.models.solde import SoldeEmploye
from app.models.user import Utilisateur
from app.schemas.demande import DemandeOut, DemandeRejeter
from app.services.auth import require_admin, require_rh
from app.services.notifications import notifier
from app.services.pdf_generator import generate_pdf
from app.services.soldes import appliquer_deduction_sur_validation

router = APIRouter()


def _get_rh(current_user: Utilisateur, db: Session) -> RH:
    rh = db.query(RH).filter(RH.utilisateur_id == current_user.id).first()
    if not rh:
        raise HTTPException(status_code=404, detail="Profil RH introuvable")
    return rh


def _enrich_demande(d: Demande) -> dict:
    out = DemandeOut.model_validate(d).model_dump()
    if d.employe and d.employe.utilisateur:
        out["nom_employe"] = d.employe.utilisateur.nom
        out["matricule"] = d.employe.matricule
    return out


# ─── Demandes ────────────────────────────────────────────────────────────────

@router.get("/demandes")
def liste_demandes(
    statut: str | None = None,
    type: str | None = None,
    search: str | None = None,
    current_user: Utilisateur = Depends(require_rh),
    db: Session = Depends(get_db),
):
    q = db.query(Demande)
    if statut:
        q = q.filter(Demande.statut == statut)
    if type:
        q = q.filter(Demande.type == type)
    if search:
        q = q.join(Demande.employe).join(Employe.utilisateur).filter(
            Utilisateur.nom.ilike(f"%{search}%")
        )
    demandes = q.order_by(Demande.created_at.desc()).all()
    return [_enrich_demande(d) for d in demandes]


@router.get("/demandes/{demande_id}")
def detail_demande(
    demande_id: int,
    current_user: Utilisateur = Depends(require_rh),
    db: Session = Depends(get_db),
):
    demande = db.query(Demande).filter(Demande.id == demande_id).first()
    if not demande:
        raise HTTPException(status_code=404, detail="Demande introuvable")
    return _enrich_demande(demande)


class UpdateDonneesPayload(BaseModel):
    donnees_collectees: dict


@router.get("/demandes/{demande_id}/preview")
def preview_demande(
    demande_id: int,
    current_user: Utilisateur = Depends(require_rh),
    db: Session = Depends(get_db),
):
    from app.services.pdf_generator import render_html
    demande = db.query(Demande).filter(Demande.id == demande_id).first()
    if not demande:
        raise HTTPException(status_code=404, detail="Demande introuvable")
    try:
        html, donnees = render_html(db, demande_id)
        # Convert all values to string for the frontend form
        donnees_str = {k: str(v) for k, v in donnees.items()}
        return {"html": html, "donnees": donnees_str}
    except (ValueError, RuntimeError) as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/demandes/{demande_id}/donnees")
def update_donnees(
    demande_id: int,
    payload: UpdateDonneesPayload,
    current_user: Utilisateur = Depends(require_rh),
    db: Session = Depends(get_db),
):
    demande = db.query(Demande).filter(Demande.id == demande_id).first()
    if not demande:
        raise HTTPException(status_code=404, detail="Demande introuvable")
    if demande.statut not in ("en_attente", "en_cours"):
        raise HTTPException(status_code=400, detail="Impossible de modifier une demande déjà traitée")
    demande.donnees_collectees = payload.donnees_collectees
    flag_modified(demande, "donnees_collectees")
    db.commit()
    return {"message": "Données mises à jour"}


@router.post("/demandes/{demande_id}/valider")
def valider_demande(
    demande_id: int,
    current_user: Utilisateur = Depends(require_rh),
    db: Session = Depends(get_db),
):
    demande = db.query(Demande).filter(Demande.id == demande_id).first()
    if not demande:
        raise HTTPException(status_code=404, detail="Demande introuvable")
    if demande.statut not in ("en_attente", "en_cours"):
        raise HTTPException(status_code=400, detail=f"La demande a déjà le statut : {demande.statut}")
    # Conflit d'intérêt : un RH/admin salarié ne valide pas sa propre demande
    if demande.employe and demande.employe.utilisateur_id == current_user.id:
        raise HTTPException(status_code=403, detail="Vous ne pouvez pas valider votre propre demande")

    rh = _get_rh(current_user, db)

    try:
        document = generate_pdf(db=db, demande_id=demande_id, rh_id=rh.id)
    except (ValueError, RuntimeError) as e:
        raise HTTPException(status_code=500, detail=str(e))

    # Déduction automatique du solde correspondant
    try:
        appliquer_deduction_sur_validation(db, demande, rh_id=rh.id)
    except Exception:
        pass  # ne pas bloquer la validation si erreur de solde

    # Notifier l'employé
    if demande.employe and demande.employe.utilisateur:
        notifier(
            db=db,
            utilisateur_id=demande.employe.utilisateur.id,
            type="demande_validee",
            titre="Demande validée",
            message=f"Votre demande #{demande.id} a été validée. Le document est disponible.",
            lien=f"/employe/mes-demandes",
        )

    return {
        "message": "Demande validée et document généré avec succès",
        "document_id": document.id,
        "chemin_fichier": document.chemin_fichier,
        "demande_id": demande_id,
    }


@router.post("/demandes/{demande_id}/rejeter")
def rejeter_demande(
    demande_id: int,
    payload: DemandeRejeter,
    current_user: Utilisateur = Depends(require_rh),
    db: Session = Depends(get_db),
):
    demande = db.query(Demande).filter(Demande.id == demande_id).first()
    if not demande:
        raise HTTPException(status_code=404, detail="Demande introuvable")
    if demande.statut not in ("en_attente", "en_cours"):
        raise HTTPException(status_code=400, detail=f"La demande a déjà le statut : {demande.statut}")
    if demande.employe and demande.employe.utilisateur_id == current_user.id:
        raise HTTPException(status_code=403, detail="Vous ne pouvez pas rejeter votre propre demande")

    demande.statut = "rejetee"
    demande.raison_rejet = payload.raison
    db.commit()

    # Notifier l'employé
    if demande.employe and demande.employe.utilisateur:
        notifier(
            db=db,
            utilisateur_id=demande.employe.utilisateur.id,
            type="demande_rejetee",
            titre="Demande rejetée",
            message=f"Votre demande #{demande.id} a été rejetée. Motif : {payload.raison}",
            lien=f"/employe/mes-demandes",
        )

    return {"message": "Demande rejetée", "demande_id": demande_id}


@router.get("/stats")
def stats(
    current_user: Utilisateur = Depends(require_rh),
    db: Session = Depends(get_db),
):
    rows = db.query(Demande.type, Demande.statut, func.count(Demande.id)).group_by(Demande.type, Demande.statut).all()
    result: dict = {}
    for type_, statut, count in rows:
        result.setdefault(type_, {})[statut] = count
    totals = db.query(Demande.statut, func.count(Demande.id)).group_by(Demande.statut).all()
    return {"par_type": result, "totaux": {s: c for s, c in totals}}


@router.get("/dashboard")
def dashboard(
    current_user: Utilisateur = Depends(require_rh),
    db: Session = Depends(get_db),
):
    from app.models.document import Document as DocumentModel

    now = datetime.now(timezone.utc)
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    urgent_threshold = now - timedelta(days=2)

    # Pending demandes enriched
    pending = db.query(Demande).filter(Demande.statut == "en_attente").order_by(Demande.created_at.asc()).all()
    pending_data = []
    for d in pending:
        nom = d.employe.utilisateur.nom if d.employe and d.employe.utilisateur else "—"
        matricule = d.employe.matricule if d.employe else "—"
        poste = d.employe.poste if d.employe else "—"
        created = d.created_at.replace(tzinfo=timezone.utc) if d.created_at.tzinfo is None else d.created_at
        days_waiting = (now - created).days
        pending_data.append({
            "id": d.id,
            "type": d.type,
            "nom_employe": nom,
            "matricule": matricule,
            "poste": poste,
            "employe_id": d.employe_id,
            "donnees_collectees": d.donnees_collectees,
            "days_waiting": days_waiting,
            "urgent": days_waiting >= 2,
            "created_at": d.created_at.isoformat(),
        })

    # Recent activity (last 10 validated or rejected)
    recent = (
        db.query(Demande)
        .filter(Demande.statut.in_(["validee", "rejetee"]))
        .order_by(Demande.updated_at.desc())
        .limit(8)
        .all()
    )
    recent_data = []
    for d in recent:
        nom = d.employe.utilisateur.nom if d.employe and d.employe.utilisateur else "—"
        recent_data.append({
            "id": d.id,
            "type": d.type,
            "statut": d.statut,
            "nom_employe": nom,
            "updated_at": d.updated_at.isoformat(),
        })

    # Counts
    total_pending = len(pending_data)
    urgent_count = sum(1 for d in pending_data if d["urgent"])

    validated_today = db.query(func.count(Demande.id)).filter(
        Demande.statut == "validee",
        Demande.updated_at >= today_start.replace(tzinfo=None),
    ).scalar() or 0

    rejected_today = db.query(func.count(Demande.id)).filter(
        Demande.statut == "rejetee",
        Demande.updated_at >= today_start.replace(tzinfo=None),
    ).scalar() or 0

    total_employes = db.query(func.count(Employe.id)).scalar() or 0
    total_docs = db.query(func.count(DocumentModel.id)).scalar() or 0

    # Top demande types
    type_rows = db.query(Demande.type, func.count(Demande.id).label("n")).group_by(Demande.type).order_by(func.count(Demande.id).desc()).limit(4).all()
    top_types = [{"type": t, "count": n} for t, n in type_rows]

    return {
        "pending": pending_data,
        "recent_activity": recent_data,
        "stats": {
            "total_pending": total_pending,
            "urgent_count": urgent_count,
            "validated_today": validated_today,
            "rejected_today": rejected_today,
            "total_employes": total_employes,
            "total_docs": total_docs,
        },
        "top_types": top_types,
    }


# ─── Registre des congés ─────────────────────────────────────────────────────

def _normalize_date_to_iso(d: str) -> str:
    """Convert dd/mm/yyyy to yyyy-mm-dd; pass through if already ISO."""
    if not d:
        return d
    if "/" in d:
        parts = d.split("/")
        if len(parts) == 3 and len(parts[2]) == 4:
            return f"{parts[2]}-{parts[1]}-{parts[0]}"
    return d


def _build_conge_rows(db: Session) -> list[dict]:
    conge_types = ("demande_conge", "attestation_conge")
    demandes = (
        db.query(Demande)
        .filter(Demande.type.in_(conge_types))
        .order_by(Demande.created_at.desc())
        .all()
    )
    rows = []
    for d in demandes:
        emp = d.employe
        if not emp or not emp.utilisateur:
            continue
        dc = d.donnees_collectees or {}
        nom_complet = emp.utilisateur.nom.split(" ", 1)
        nom = nom_complet[0] if len(nom_complet) >= 1 else ""
        prenom = nom_complet[1] if len(nom_complet) >= 2 else ""
        type_conge = dc.get("type_conge", "Congé annuel")
        date_debut = dc.get("date_debut", "")
        date_fin = dc.get("date_fin", "")

        nb_jours = 0
        try:
            from app.services.soldes import _compter_jours
            if date_debut and date_fin:
                nb_jours = _compter_jours(
                    _normalize_date_to_iso(date_debut),
                    _normalize_date_to_iso(date_fin),
                )
        except Exception:
            nb_jours = 0

        solde_restant = None
        solde = db.query(SoldeEmploye).filter(
            SoldeEmploye.employe_id == emp.id,
            SoldeEmploye.type == "conge_annuel",
            SoldeEmploye.annee_reference == datetime.now().year,
        ).first()
        if solde:
            solde_restant = float(solde.quota_total) - float(solde.consomme)

        rows.append({
            "demande_id": d.id,
            "matricule": emp.matricule,
            "nom": nom,
            "prenom": prenom,
            "service": emp.departement,
            "date_embauche": emp.date_embauche.strftime("%d/%m/%Y"),
            "type_conge": type_conge,
            "date_debut": date_debut,
            "date_fin": date_fin,
            "nb_jours": nb_jours,
            "statut": d.statut.capitalize().replace("_", " "),
            "solde_restant": solde_restant,
        })
    return rows


@router.get("/registre-conges")
def registre_conges(
    current_user: Utilisateur = Depends(require_rh),
    db: Session = Depends(get_db),
):
    return _build_conge_rows(db)


@router.get("/registre-conges/export")
def export_registre_conges(
    current_user: Utilisateur = Depends(require_rh),
    db: Session = Depends(get_db),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    rows = _build_conge_rows(db)

    wb = Workbook()
    ws = wb.active
    ws.title = "Registre des Congés"

    headers = [
        "Matricule", "Nom", "Prénom", "Service", "Date d'embauche",
        "Type de congé", "Date début", "Date fin", "Nb jours", "Statut", "Solde restant",
    ]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, row in enumerate(rows, 2):
        values = [
            row["matricule"], row["nom"], row["prenom"], row["service"],
            row["date_embauche"], row["type_conge"], row["date_debut"],
            row["date_fin"], row["nb_jours"], row["statut"],
            row["solde_restant"] if row["solde_restant"] is not None else "—",
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    col_widths = [12, 15, 15, 18, 16, 18, 14, 14, 10, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"registre_conges_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ─── Employés — lecture seule (fiche détaillée) ──────────────────────────────
# ⚠️ La création / édition / suppression des employés est centralisée dans
# `users.py` (/api/users) : email & matricule auto, prénom, garde-fous, et
# suppression protégée contre les FK. On ne conserve ici que les endpoints de
# consultation utilisés par la fiche employé (/rh/employes/[id]).


def _employe_to_dict(e: Employe) -> dict:
    return {
        "id": e.id,
        "utilisateur_id": e.utilisateur_id,
        "nom": e.utilisateur.nom,
        "email": e.utilisateur.email,
        "matricule": e.matricule,
        "poste": e.poste,
        "departement": e.departement,
        "salaire_base": float(e.salaire_base),
        "date_embauche": e.date_embauche.isoformat(),
        "statut": e.statut,
        "type_contrat": e.type_contrat,
        "cin": e.cin,
        "cnss": e.cnss,
        "adresse": e.adresse,
        "telephone": e.telephone,
    }


@router.get("/employes/{employe_id}")
def get_employe(
    employe_id: int,
    current_user: Utilisateur = Depends(require_rh),
    db: Session = Depends(get_db),
):
    e = db.query(Employe).filter(Employe.id == employe_id).first()
    if not e:
        raise HTTPException(status_code=404, detail="Employé introuvable")
    return _employe_to_dict(e)


@router.get("/employes/{employe_id}/demandes")
def get_employe_demandes(
    employe_id: int,
    current_user: Utilisateur = Depends(require_rh),
    db: Session = Depends(get_db),
):
    e = db.query(Employe).filter(Employe.id == employe_id).first()
    if not e:
        raise HTTPException(status_code=404, detail="Employé introuvable")
    demandes = db.query(Demande).filter(Demande.employe_id == employe_id).order_by(Demande.created_at.desc()).all()
    return [_enrich_demande(d) for d in demandes]


# ─── Paramétrage (signature + cachet) ───────────────────────────────────────

def _param_image_data_uri(valeur: str | None) -> str | None:
    """Lit le fichier signature/cachet et le renvoie en data URI base64.

    L'aperçu RH ne dépend plus d'un dossier `/uploads` public (retiré pour
    raisons de sécurité) : l'image transite dans la réponse JSON protégée.
    `valeur` est le chemin FS stocké (ex. « /uploads/parametrage/signature.png »).
    """
    import base64
    import os

    if not valeur:
        return None
    abs_path = os.path.join(".", valeur.lstrip("/"))
    if not os.path.exists(abs_path):
        return None
    ext = abs_path.rsplit(".", 1)[-1].lower() if "." in abs_path else "png"
    mime = {"jpg": "jpeg", "jpeg": "jpeg", "png": "png", "gif": "gif", "svg": "svg+xml", "webp": "webp"}.get(ext, "png")
    with open(abs_path, "rb") as f:
        data = base64.b64encode(f.read()).decode("ascii")
    return f"data:image/{mime};base64,{data}"


@router.get("/parametrage")
def get_parametrage(
    current_user: Utilisateur = Depends(require_admin),
    db: Session = Depends(get_db),
):
    sig = db.query(Parametrage).filter(Parametrage.cle == "signature").first()
    cachet = db.query(Parametrage).filter(Parametrage.cle == "cachet").first()
    return {
        "signature_url": _param_image_data_uri(sig.valeur if sig else None),
        "cachet_url": _param_image_data_uri(cachet.valeur if cachet else None),
    }


def _save_parametrage_image(db: Session, cle: str, file: UploadFile) -> str:
    import os
    upload_dir = os.path.join(settings.UPLOADS_DIR, "parametrage")
    os.makedirs(upload_dir, exist_ok=True)

    ext = file.filename.rsplit(".", 1)[-1] if "." in file.filename else "png"
    filename = f"{cle}.{ext}"
    filepath = os.path.join(upload_dir, filename)

    with open(filepath, "wb") as f:
        f.write(file.file.read())

    url_path = f"/uploads/parametrage/{filename}"

    row = db.query(Parametrage).filter(Parametrage.cle == cle).first()
    if row:
        row.valeur = url_path
    else:
        db.add(Parametrage(cle=cle, valeur=url_path))
    db.commit()
    return url_path


@router.post("/parametrage/signature")
def upload_signature(
    file: UploadFile = FastAPIFile(...),
    current_user: Utilisateur = Depends(require_admin),
    db: Session = Depends(get_db),
):
    url = _save_parametrage_image(db, "signature", file)
    return {"signature_url": _param_image_data_uri(url)}


@router.post("/parametrage/cachet")
def upload_cachet(
    file: UploadFile = FastAPIFile(...),
    current_user: Utilisateur = Depends(require_admin),
    db: Session = Depends(get_db),
):
    url = _save_parametrage_image(db, "cachet", file)
    return {"cachet_url": _param_image_data_uri(url)}


@router.delete("/parametrage/signature")
def delete_signature(
    current_user: Utilisateur = Depends(require_admin),
    db: Session = Depends(get_db),
):
    row = db.query(Parametrage).filter(Parametrage.cle == "signature").first()
    if row:
        import os
        filepath = os.path.join(".", row.valeur.lstrip("/"))
        if os.path.exists(filepath):
            os.remove(filepath)
        db.delete(row)
        db.commit()
    return {"message": "Signature supprimée"}


@router.delete("/parametrage/cachet")
def delete_cachet(
    current_user: Utilisateur = Depends(require_admin),
    db: Session = Depends(get_db),
):
    row = db.query(Parametrage).filter(Parametrage.cle == "cachet").first()
    if row:
        import os
        filepath = os.path.join(".", row.valeur.lstrip("/"))
        if os.path.exists(filepath):
            os.remove(filepath)
        db.delete(row)
        db.commit()
    return {"message": "Cachet supprimé"}


# ─── Paramétrage IA (chatbot Groq + embeddings RAG) ─────────────────────────

class IAParams(BaseModel):
    groq_api_key: Optional[str] = None
    groq_api_key_2: Optional[str] = None
    groq_model: Optional[str] = None
    embedding_provider: Optional[str] = None
    hf_api_key: Optional[str] = None
    hf_embedding_model: Optional[str] = None
    ollama_base_url: Optional[str] = None
    ollama_embedding_model: Optional[str] = None


@router.get("/parametrage/ia")
def get_parametrage_ia(
    current_user: Utilisateur = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """État de la config IA — les clés secrètes ne sont JAMAIS renvoyées (juste configurée ou non)."""
    from app.services.parametrage import get_param
    return {
        "groq_api_key_set": bool(get_param(db, "groq_api_key", settings.GROQ_API_KEY)),
        "groq_api_key_2_set": bool(get_param(db, "groq_api_key_2", "")),
        "groq_model": get_param(db, "groq_model", settings.GROQ_MODEL),
        "embedding_provider": get_param(db, "embedding_provider", settings.EMBEDDING_PROVIDER),
        "hf_api_key_set": bool(get_param(db, "hf_api_key", settings.HF_API_KEY)),
        "hf_embedding_model": get_param(db, "hf_embedding_model", settings.HF_EMBEDDING_MODEL),
        "ollama_base_url": get_param(db, "ollama_base_url", settings.OLLAMA_BASE_URL),
        "ollama_embedding_model": get_param(db, "ollama_embedding_model", settings.OLLAMA_EMBEDDING_MODEL),
    }


@router.post("/parametrage/ia")
def save_parametrage_ia(
    body: IAParams,
    current_user: Utilisateur = Depends(require_admin),
    db: Session = Depends(get_db),
):
    from app.services.parametrage import set_param
    data = body.model_dump(exclude_none=True)
    changed_embed = False
    for cle, val in data.items():
        val = str(val)
        # Ne pas écraser une clé secrète par une valeur vide (permet de la conserver)
        if cle in ("groq_api_key", "groq_api_key_2", "hf_api_key") and not val.strip():
            continue
        set_param(db, cle, val)
        if cle.startswith(("embedding", "hf_", "ollama")):
            changed_embed = True
    db.commit()
    if changed_embed:
        # Provider/paramètres d'embeddings modifiés → forcer la reconstruction du RAG
        from app.services.rag import reset_rag_service
        reset_rag_service()
    return {"message": "Paramètres IA enregistrés"}


@router.post("/parametrage/ia/test-groq")
def test_parametrage_groq(
    current_user: Utilisateur = Depends(require_admin),
    db: Session = Depends(get_db),
):
    from app.services.parametrage import groq_keys, groq_model
    from langchain_groq import ChatGroq
    keys = groq_keys(db)
    if not keys:
        raise HTTPException(status_code=400, detail="Aucune clé Groq configurée")
    try:
        llm = ChatGroq(model=groq_model(db), api_key=keys[0], temperature=0)
        llm.invoke("ping")
        return {"ok": True, "model": groq_model(db), "cles": len(keys)}
    except Exception as e:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=f"Échec Groq : {e}")


# ─── Paramétrage SMTP (invitations par email) ───────────────────────────────

class SMTPParams(BaseModel):
    smtp_host: Optional[str] = None
    smtp_port: Optional[str] = None
    smtp_user: Optional[str] = None
    smtp_password: Optional[str] = None
    smtp_from: Optional[str] = None


@router.get("/parametrage/smtp")
def get_parametrage_smtp(
    current_user: Utilisateur = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """État SMTP — le mot de passe n'est jamais renvoyé (juste configuré ou non)."""
    from app.services.parametrage import get_param
    return {
        "smtp_host": get_param(db, "smtp_host", settings.SMTP_HOST),
        "smtp_port": get_param(db, "smtp_port", str(settings.SMTP_PORT)),
        "smtp_user": get_param(db, "smtp_user", settings.SMTP_USER),
        "smtp_from": get_param(db, "smtp_from", settings.SMTP_FROM),
        "smtp_password_set": bool(get_param(db, "smtp_password", settings.SMTP_PASSWORD)),
    }


@router.post("/parametrage/smtp")
def save_parametrage_smtp(
    body: SMTPParams,
    current_user: Utilisateur = Depends(require_admin),
    db: Session = Depends(get_db),
):
    from app.services.parametrage import set_param
    data = body.model_dump(exclude_none=True)
    for cle, val in data.items():
        val = str(val)
        # Ne pas écraser le mot de passe par une valeur vide (permet de le conserver)
        if cle == "smtp_password" and not val.strip():
            continue
        set_param(db, cle, val)
    db.commit()
    return {"message": "Paramètres SMTP enregistrés"}
